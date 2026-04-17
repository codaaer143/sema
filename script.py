import pandas as pd
import datetime
import io
import openpyxl
import os
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

# =========================
# FOLDER PATHS
# =========================
INPUT_FOLDER = "Input.1"
OUTPUT_FOLDER = "Output.2"

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# =========================
# PROCESS FILES
# =========================
for file_name in os.listdir(INPUT_FOLDER):
    if file_name.endswith(".xlsx") or file_name.endswith(".xls"):

        file_path = os.path.join(INPUT_FOLDER, file_name)

        # =========================
        # LOAD RAW DATA
        # =========================
        df_raw = pd.read_excel(file_path)
        raw_rows = len(df_raw)

        target_columns = ['Material', 'Description', 'R/D Date', 'R/D Time', 'DO Qty', 'D/L', 'Line']
        column_names = ['Business', 'Material', 'Description', 'R/D Date', 'R/D Time', 'DO Qty', 'D/L', 'Line']

        if all(col in df_raw.columns for col in column_names):

            # =========================
            # CLEAN
            # =========================
            df = df_raw.drop(columns=[col for col in df_raw.columns if col not in target_columns], errors='ignore')

            df = df.rename(columns={
                'Material': 'Part Number',
                'R/D Date': 'Date',
                'R/D Time': 'Time',
                'DO Qty': 'Qty',
                'D/L': 'Location',
            })

            # =========================
            # FORMAT FIRST (CRITICAL FIX)
            # =========================
            df['Date'] = pd.to_datetime(df['Date'], format='%d/%m/%Y', errors='coerce')
            df['Time'] = pd.to_datetime(df['Time'], format='%H:%M:%S', errors='coerce')

            df['Time'] = df['Time'].dt.strftime('%I:%M %p')
            df['Date'] = df['Date'].dt.date

            # =========================
            # CLEAN LINE SAFELY
            # =========================
            df['Line'] = df['Line'].astype(str).str.strip().str.upper()

            df.loc[
                (df['Line'].isna()) |
                (df['Line'] == '') |
                (df['Line'] == 'NAN') |
                (df['Line'] == 'NONE'),
                'Line'
            ] = df['Location']

            # =========================
            # CLEAN ROWS (SAFETY)
            # =========================
            df = df.dropna(subset=['Date', 'Time', 'Qty'])

            cleaned_rows = len(df)

            # =========================
            # GROUP
            # =========================
            df_grouped = df.groupby(
                ['Date', 'Part Number', 'Description', 'Location', 'Time', 'Line'],
                as_index=False
            )['Qty'].sum()

            grouped_rows = len(df_grouped)

            df_grouped = df_grouped.sort_values(['Date', 'Time', 'Location'])

            # =========================
            # REORDER
            # =========================
            df_final = df_grouped[['Date', 'Location', 'Time', 'Description', 'Part Number', 'Qty']]

            # =========================
            # EXPORT EXCEL
            # =========================
            output_buffer = io.BytesIO()

            with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                df_final.to_excel(writer, index=False, sheet_name="Weekly Delivery Order")

                # =========================
                # DATA INTEGRITY SHEET (NEW)
                # =========================
                check_df = pd.DataFrame({
                    "Metric": [
                        "Raw Rows",
                        "After Cleaning",
                        "After Grouping",
                        "Possible Compression Ratio"
                    ],
                    "Value": [
                        raw_rows,
                        cleaned_rows,
                        grouped_rows,
                        round(grouped_rows / raw_rows, 2) if raw_rows else 0
                    ]
                })

                check_df.to_excel(writer, index=False, sheet_name="DATA_CHECK")

            workbook = openpyxl.load_workbook(output_buffer)
            sheet = workbook["Weekly Delivery Order"]

            # =========================
            # INSERT HEADER SPACE
            # =========================
            sheet.insert_rows(1, amount=2)

            # =========================
            # DATE HEADER
            # =========================
            unique_dates = sorted(df_final['Date'].unique())
            formatted_dates = [pd.to_datetime(d).strftime('%d.%m.%Y') for d in unique_dates]

            sheet["A1"] = "DATE : " + " & ".join(formatted_dates)

            sheet["A1"].font = Font(name="Book Antiqua", bold=True, size=16)
            sheet["A1"].alignment = Alignment(horizontal='left')

            # =========================
            # GLOBAL FONT
            # =========================
            global_font = Font(name="Book Antiqua")

            for row in sheet.iter_rows():
                for cell in row:
                    cell.font = global_font

            # =========================
            # HEADER STYLE
            # =========================
            header_font = Font(name="Book Antiqua", bold=True)
            for cell in sheet[3]:
                cell.font = header_font

            # =========================
            # BORDER
            # =========================
            def set_border(ws, cell_range, border_style):
                rows = cell_range.split(':')
                start = rows[0]
                end = rows[1] if len(rows) > 1 else rows[0]

                border = Border(
                    left=border_style,
                    right=border_style,
                    top=border_style,
                    bottom=border_style
                )

                for row in ws[start:end]:
                    for cell in row:
                        if cell.value is not None:
                            cell.border = border

            set_border(sheet, f"A3:F{sheet.max_row}", Side(border_style='thin', color='000000'))

            # =========================
            # RC12 FORMAT (SAFE)
            # =========================
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
                location_cell = row[1]

                if str(location_cell.value).strip().upper() == "RC12":
                    for cell in row:
                        cell.fill = yellow_fill
                        cell.font = Font(name="Book Antiqua")

            # =========================
            # ALIGNMENT
            # =========================
            for col in ['A', 'B', 'C', 'E', 'F']:
                for cell in sheet[col]:
                    cell.alignment = Alignment(horizontal='center')

            for cell in sheet['D']:
                cell.alignment = Alignment(horizontal='left')

            # =========================
            # AUTO FIT (B-F ONLY)
            # =========================
            sheet.column_dimensions['A'].width = 18

            for col in sheet.columns:
                col_letter = col[0].column_letter

                if col_letter == 'A':
                    continue

                max_length = max(
                    (len(str(cell.value)) for cell in col if cell.value),
                    default=10
                )

                sheet.column_dimensions[col_letter].width = max_length + 3

            # =========================
            # FILTER
            # =========================
            sheet.auto_filter.ref = f"A3:F{sheet.max_row}"

            # =========================
            # SAVE OUTPUT
            # =========================
            today = datetime.datetime.now().strftime("%m-%d-%Y")
            output_file_name = f"{today} Weekly Delivery Order_{file_name}"

            output_path = os.path.join(OUTPUT_FOLDER, output_file_name)
            workbook.save(output_path)

            print(f"Processed: {file_name} -> {output_file_name}")

        else:
            print(f"SKIPPED (wrong columns): {file_name}")
