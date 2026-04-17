import pandas as pd
import datetime
import io
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill
from flask import Flask, request, send_file, render_template_string

# =========================
# FLASK APP INITIALIZATION
# =========================
app = Flask(__name__)

# =========================
# HTML TEMPLATE FOR UPLOAD
# =========================
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Excel File Processor</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 40px; }
        h1 { color: #333; }
        .container { max-width: 500px; margin: auto; padding: 20px; border: 1px solid #ddd; border-radius: 5px; }
        .btn { background-color: #4CAF50; color: white; padding: 10px 15px; border: none; border-radius: 4px; cursor: pointer; }
        .btn:hover { background-color: #45a049; }
    </style>
</head>
<body>
    <div class="container">
        <h1>Upload Excel File</h1>
        <form method="post" enctype="multipart/form-data">
            <input type="file" name="file" accept=".xlsx, .xls">
            <button type="submit" class="btn">Process File</button>
        </form>
    </div>
</body>
</html>
"""

# =========================
# DATA PROCESSING FUNCTION
# =========================
def process_excel_file(input_file, original_filename):
    """
    Processes the uploaded Excel file in memory and returns the processed file as a BytesIO buffer.
    """
    # =========================
    # LOAD RAW DATA
    # =========================
    df_raw = pd.read_excel(input_file)
    raw_rows = len(df_raw)

    target_columns = ['Material', 'Description', 'R/D Date', 'R/D Time', 'DO Qty', 'D/L', 'Line']
    column_names = ['Business', 'Material', 'Description', 'R/D Date', 'R/D Time', 'DO Qty', 'D/L', 'Line']

    if not all(col in df_raw.columns for col in column_names):
        raise ValueError(f"SKIPPED (wrong columns): {original_filename}")

    # =========================
    # CLEAN
    # =========================
    df = df_raw.drop(columns=[col for col in df_raw.columns if col not in target_columns], errors='ignore')

    df = df.rename(columns={
        'Material': 'Part Number',
        'R/D Date': 'Date',
        'R/D Time': 'Time',
        'DO Qty': 'Qty',
        'D/L': 'Location',
    })

    # =========================
    # FORMAT FIRST (CRITICAL FIX)
    # =========================
    df['Date'] = pd.to_datetime(df['Date'], format='%d/%m/%Y', errors='coerce')
    df['Time'] = pd.to_datetime(df['Time'], format='%H:%M:%S', errors='coerce')

    df['Time'] = df['Time'].dt.strftime('%I:%M %p')
    df['Date'] = df['Date'].dt.date

    # =========================
    # CLEAN LINE SAFELY
    # =========================
    df['Line'] = df['Line'].astype(str).str.strip().str.upper()
    df.loc[
        (df['Line'].isna()) | (df['Line'] == '') | (df['Line'] == 'NAN') | (df['Line'] == 'NONE'),
        'Line'
    ] = df['Location']

    # =========================
    # CLEAN ROWS (SAFETY)
    # =========================
    df = df.dropna(subset=['Date', 'Time', 'Qty'])
    cleaned_rows = len(df)

    # =========================
    # GROUP
    # =========================
    df_grouped = df.groupby(
        ['Date', 'Part Number', 'Description', 'Location', 'Time', 'Line'],
        as_index=False
    )['Qty'].sum()
    grouped_rows = len(df_grouped)
    df_grouped = df_grouped.sort_values(['Date', 'Time', 'Location'])

    # =========================
    # REORDER
    # =========================
    df_final = df_grouped[['Date', 'Location', 'Time', 'Description', 'Part Number', 'Qty']]

    # =========================
    # EXPORT EXCEL
    # =========================
    output_buffer = io.BytesIO()
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, sheet_name="Weekly Delivery Order")

        # DATA INTEGRITY SHEET
        check_df = pd.DataFrame({
            "Metric": ["Raw Rows", "After Cleaning", "After Grouping", "Possible Compression Ratio"],
            "Value": [raw_rows, cleaned_rows, grouped_rows, round(grouped_rows / raw_rows, 2) if raw_rows else 0]
        })
        check_df.to_excel(writer, index=False, sheet_name="DATA_CHECK")

    workbook = openpyxl.load_workbook(output_buffer)
    sheet = workbook["Weekly Delivery Order"]

    # =========================
    # STYLING
    # =========================
    sheet.insert_rows(1, amount=2)
    unique_dates = sorted(df_final['Date'].unique())
    formatted_dates = [pd.to_datetime(d).strftime('%d.%m.%Y') for d in unique_dates]
    sheet["A1"] = "DATE : " + " & ".join(formatted_dates)
    sheet["A1"].font = Font(name="Book Antiqua", bold=True, size=16)
    sheet["A1"].alignment = Alignment(horizontal='left')

    global_font = Font(name="Book Antiqua")
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = global_font

    header_font = Font(name="Book Antiqua", bold=True)
    for cell in sheet[3]:
        cell.font = header_font

    thin_border_side = Side(border_style='thin', color='000000')
    border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    for row in sheet[f"A3:F{sheet.max_row}"]:
        for cell in row:
            if cell.value is not None:
                cell.border = border

    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row):
        location_cell = row[1]
        if str(location_cell.value).strip().upper() == "RC12":
            for cell in row:
                cell.fill = yellow_fill

    for col_letter in ['A', 'B', 'C', 'E', 'F']:
        for cell in sheet[col_letter]:
            cell.alignment = Alignment(horizontal='center')
    for cell in sheet['D']:
        cell.alignment = Alignment(horizontal='left')

    sheet.column_dimensions['A'].width = 18
    for col in sheet.columns:
        col_letter = col[0].column_letter
        if col_letter != 'A':
            max_length = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            sheet.column_dimensions[col_letter].width = max_length + 3

    sheet.auto_filter.ref = f"A3:F{sheet.max_row}"

    # =========================
    # SAVE TO BUFFER
    # =========================
    final_output_buffer = io.BytesIO()
    workbook.save(final_output_buffer)
    final_output_buffer.seek(0)

    # =========================
    # GENERATE FILENAME
    # =========================
    today = datetime.datetime.now().strftime("%m-%d-%Y")
    output_file_name = f"{today} Weekly Delivery Order_{original_filename}"

    return final_output_buffer, output_file_name

# =========================
# FLASK ROUTES
# =========================
@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return 'No file part'
        file = request.files['file']
        if file.filename == '':
            return 'No selected file'
        if file and (file.filename.endswith(".xlsx") or file.filename.endswith(".xls")):
            try:
                processed_buffer, output_filename = process_excel_file(file, file.filename)
                return send_file(
                    processed_buffer,
                    as_attachment=True,
                    download_name=output_filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            except Exception as e:
                return str(e)
        else:
            return "Invalid file type. Please upload an .xlsx or .xls file."

    return render_template_string(HTML_TEMPLATE)

# =========================
# RUN THE APP
# =========================
if __name__ == '__main__':
    app.run(debug=True)

